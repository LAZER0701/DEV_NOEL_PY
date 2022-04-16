from openpyxl import load_workbook, Workbook

c_name = 1
c_id = 2
c_money = 3
c_lvl = 4

default_money = 10000

wb = load_workbook("userDB.xlsx")
ws = wb.active

def checkRow():
    for row in range(2, ws.max_row + 1):
        if ws.cell(row,1).value is None:
            return row
            break

def signup(_name, _id):
    _row = checkRow()
    
    ws.cell(row=_row, column=c_name, value=_name)
    ws.cell(row=_row, column=c_id, value =_id)
    ws.cell(row=_row, column=c_money, value = default_money)
    ws.cell(row=_row, column=c_lvl, value = 1)

def checkName(_name, _id):
    for row in range(2, ws.max_row+1):
        if ws.cell(row,1).value == _name and ws.cell(row,2).value == _id:
            break
            return False
        else:
            return True
            break

def userInfo(_namd, _id):
    	if not checkName(_namd, _id):
    	for row in range(2, ws.max_row+2):
    		if ws.cell(row, 1).value == _name and ws.cell(row, 2).value == _id:
        		return ws.cell(row,1).value, ws.cell(row,c_lvl).value
                break
    else:
    	return None, None
        
    wb.save("userDB.xlsx")